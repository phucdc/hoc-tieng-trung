from flask import Flask, request, jsonify, render_template, url_for, send_file
import openpyxl
import cruds.tab as t_crud
import cruds.word as w_crud
from database.database import init_db
from config import APP_DIR, SAVE_VOICE_DIR
import os
import random

app = Flask(__name__, template_folder='templates', static_folder='static')
app.config['TEMPLATES_AUTO_RELOAD'] = True


@app.context_processor
def storage_processor():
    def storage_url(path):
        storage_dir = os.path.join(app.root_path, 'storage')
        return url_for('storage', filename=path, _external=True, _scheme='http')
    return dict(storage=storage_url)


@app.route('/storage/<path:filename>')
def storage(filename):
    return send_file(os.path.join(SAVE_VOICE_DIR, filename))


@app.route("/")
async def index():
    words = w_crud.get_all_words()
    tabs = t_crud.get_all_tabs()
    index = render_template('index.html', words=words, tabs=tabs)
    return index
    

@app.post('/word/add')
async def add_word():
    word = request.json.get('word')
    pinyin = request.json.get('pinyin')
    meaning = request.json.get('meaning')
    w = w_crud.create(word, pinyin, meaning)
    words = w_crud.get_all_words()
    tabs = t_crud.get_all_tabs()
    index = render_template('main-section.html', words=words, tabs=tabs)
    return index


@app.post('/learn')
async def learn():
    print(request.form)
    word_ids = request.json
    words = []
    for id in word_ids:
        word = w_crud.read(id)
        if word:
            words.append(word)
    tabs = t_crud.get_all_tabs()
    return render_template('learn.html', words=words, tabs=tabs)


@app.post('/tab/add_word')
async def add_to_tab():
    print(request.json)
    word_ids = request.json['word_ids']
    tab_id = request.json['tab_id']
    for id in word_ids:
        t_crud.add_word_to_tab(word_id=id, tab_id=tab_id)
    return tab_id


@app.post('/tab/get_words')
async def get_words_from_tab():
    tab_id = request.json['tab_id']
    res = []
    for word in t_crud.get_words_from_tab(tab_id=tab_id):
        res.append(word.__to_dict__())
    tabs = t_crud.get_all_tabs()
    return render_template('learn.html', words=res, tabs=tabs)


@app.post('/tab/add')
async def add_tab():
    name = request.json['tab_name']
    tab = t_crud.create(name=name)
    tabs = t_crud.get_all_tabs()
    return render_template('tab.html', tabs=tabs)


@app.post("/import")
async def import_handle():
    validate_header = 'word,pinyin,meaning'
    f = request.files.get('list')
    if f:
        data = f.stream
        workbook = openpyxl.load_workbook(data, data_only=True)
        worksheet = workbook.active
        
        header_row = next(worksheet.iter_rows(min_row=1, values_only=True))
        
        if ','.join(header_row).lower() == validate_header:
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                word, pinyin, meaning = row
                w_crud.create(word=word, pinyin=pinyin, meaning=meaning)
                
            return render_template('import.html', color="success", status="Succeed")
        
    return render_template('import.html', color="danger", status="Failed")


if __name__ == '__main__':
    init_db()
    app.run(
        host='0.0.0.0',
        port=8000
    )
